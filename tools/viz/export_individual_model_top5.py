"""Export a supervisor-friendly table of separate BERT/GloVe/Word2Vec top-5s.

The workbook deliberately excludes any combined systems such as
Word2Vec+GloVe or VLM rank fusion. Each row is one video; each model has one
compact top-5 column and five separate keyword columns for easy comparison.
"""

from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_HERE = Path(__file__).resolve()
_ROOT = _HERE.parent.parent.parent

DEFAULT_REPORT = _ROOT / "reports" / "semantic_model_comparison_15videos_graded_nocache.json"
DEFAULT_GRADED = _ROOT / "data" / "gold" / "videos_15_graded_vlm_reference_nocache.json"
DEFAULT_XLSX = _ROOT / "reports" / "individual_model_top5_keywords_15videos.xlsx"
DEFAULT_CSV = _ROOT / "reports" / "individual_model_top5_keywords_15videos.csv"

MODEL_ORDER = ("bert", "glove", "word2vec")
MODEL_LABELS = {
    "bert": "BERT",
    "glove": "GloVe",
    "word2vec": "Word2Vec",
}

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
NOTE_FILL = PatternFill("solid", fgColor="EAF2F8")


def _load_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def _combo_map(report: dict[str, Any]) -> dict[str, dict[str, dict[str, Any]]]:
    combos: dict[str, dict[str, dict[str, Any]]] = {}
    for combo in report.get("combos") or []:
        model = str(combo.get("model") or "").strip().lower()
        if model not in MODEL_ORDER:
            continue
        combos[model] = {
            str(item.get("item_id") or ""): item
            for item in combo.get("items") or []
            if isinstance(item, dict)
        }
    missing = [m for m in MODEL_ORDER if m not in combos]
    if missing:
        raise ValueError(f"Missing model results in report: {', '.join(missing)}")
    return combos


def _items(graded: dict[str, Any]) -> list[dict[str, Any]]:
    items = [item for item in graded.get("items") or [] if isinstance(item, dict)]
    if not items:
        raise ValueError("No graded video items found.")
    return items


def _top_keywords(rec: dict[str, Any], k: int = 5) -> list[str]:
    raw = rec.get("top_keywords") or []
    out = [str(x).strip() for x in raw if str(x).strip()]
    return (out + [""] * k)[:k]


def _top_scores(rec: dict[str, Any], k: int = 5) -> list[str]:
    ranked = rec.get("ranking") or []
    out: list[str] = []
    for row in ranked[:k]:
        try:
            out.append(f"{float(row.get('score')):.3f}")
        except Exception:
            out.append("")
    return (out + [""] * k)[:k]


def _build_rows(report: dict[str, Any], graded: dict[str, Any]) -> list[dict[str, Any]]:
    combos = _combo_map(report)
    rows: list[dict[str, Any]] = []
    for i, item in enumerate(_items(graded), start=1):
        vid = str(item.get("id") or item.get("item_id") or "").strip()
        row: dict[str, Any] = {
            "video_no": i,
            "video_id": vid,
            "url": item.get("url") or "",
            "title": item.get("title") or "",
            "topic": item.get("topic") or "",
            "subtopic": item.get("subtopic") or "",
        }
        for model in MODEL_ORDER:
            rec = combos[model].get(vid, {})
            kws = _top_keywords(rec)
            scores = _top_scores(rec)
            label = model
            row[f"{label}_top5"] = ", ".join(kws)
            row[f"{label}_top5_with_scores"] = ", ".join(
                f"{kw} ({score})" if score else kw for kw, score in zip(kws, scores)
            )
            for j, kw in enumerate(kws, start=1):
                row[f"{label}_keyword_{j}"] = kw
            for j, score in enumerate(scores, start=1):
                row[f"{label}_score_{j}"] = score
        rows.append(row)
    return rows


def _style_header(ws) -> None:
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _autosize(ws, max_width: int = 64) -> None:
    for idx, col in enumerate(ws.columns, start=1):
        width = 10
        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            parts = value.splitlines() or [value]
            width = max(width, min(max_width, max(len(part) for part in parts)))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.column_dimensions[get_column_letter(idx)].width = width + 2


def _append_sheet(ws, headers: list[str], rows: list[dict[str, Any]]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    _style_header(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    _autosize(ws)


def export(report_path: Path, graded_path: Path, xlsx_path: Path, csv_path: Path) -> None:
    report = _load_json(report_path)
    graded = _load_json(graded_path)
    rows = _build_rows(report, graded)

    compact_headers = [
        "video_no",
        "video_id",
        "url",
        "title",
        "topic",
        "subtopic",
        "bert_top5",
        "glove_top5",
        "word2vec_top5",
    ]
    separate_headers = [
        "video_no",
        "video_id",
        "url",
        "title",
        "topic",
        "subtopic",
    ]
    for model in MODEL_ORDER:
        for j in range(1, 6):
            separate_headers.append(f"{model}_keyword_{j}")
    scored_headers = [
        "video_no",
        "video_id",
        "title",
        "bert_top5_with_scores",
        "glove_top5_with_scores",
        "word2vec_top5_with_scores",
    ]

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    csv_path.parent.mkdir(parents=True, exist_ok=True)

    with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=compact_headers)
        writer.writeheader()
        writer.writerows([{h: row.get(h, "") for h in compact_headers} for row in rows])

    wb = Workbook()
    ws = wb.active
    ws.title = "Top5 Compact"
    _append_sheet(ws, compact_headers, rows)

    ws_sep = wb.create_sheet("Top5 Separate Columns")
    _append_sheet(ws_sep, separate_headers, rows)

    ws_scores = wb.create_sheet("Top5 With Scores")
    _append_sheet(ws_scores, scored_headers, rows)

    ws_note = wb.create_sheet("Methodology")
    notes = [
        ("Purpose", "Separate top-5 keyword outputs for BERT, GloVe, and Word2Vec."),
        ("Important", "No model combination is included in these results."),
        ("Input", "Same 15 videos, fixed VLM title+summary/query text, same closed keyword pool."),
        ("Metric", "Cosine similarity."),
        ("Models", "BERT, GloVe, and Word2Vec are evaluated separately."),
        ("Output", "Each model contributes exactly five keywords per video."),
    ]
    ws_note.append(["Field", "Value"])
    for row in notes:
        ws_note.append(list(row))
    _style_header(ws_note)
    for row in ws_note.iter_rows(min_row=2):
        row[0].fill = NOTE_FILL
        row[0].font = Font(bold=True)
    _autosize(ws_note, max_width=90)

    wb.save(xlsx_path)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--report", type=Path, default=DEFAULT_REPORT)
    parser.add_argument("--graded", type=Path, default=DEFAULT_GRADED)
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    parser.add_argument("--csv", type=Path, default=DEFAULT_CSV)
    args = parser.parse_args()

    export(
        report_path=args.report.resolve(),
        graded_path=args.graded.resolve(),
        xlsx_path=args.xlsx.resolve(),
        csv_path=args.csv.resolve(),
    )
    print(f"[OK] XLSX: {args.xlsx.resolve()}")
    print(f"[OK] CSV: {args.csv.resolve()}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
