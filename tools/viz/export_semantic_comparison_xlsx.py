"""
Export semantic-model comparison results to an Excel workbook.

The workbook is designed for supervisor-facing review:
  - Summary metrics and charts
  - Per-video top-5 keyword choices for BERT/Word2Vec/GloVe
  - Long-form per-model metrics
  - Graded VLM reference keywords
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_HERE = Path(__file__).resolve()
_ROOT = _HERE.parent.parent.parent


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SUB_FILL = PatternFill("solid", fgColor="D9EAF7")


def _load(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))


def _abs(path: str) -> Path:
    p = Path(path)
    return p if p.is_absolute() else (_ROOT / p).resolve()


def _join_keywords(items) -> str:
    return "\n".join(str(x) for x in (items or []))


def _join_graded(items) -> str:
    out = []
    for row in items or []:
        if not isinstance(row, dict):
            continue
        kw = row.get("keyword")
        rel = row.get("relevance")
        if kw:
            out.append(f"{kw} ({rel})")
    return "\n".join(out)


def _fmt_float(value):
    try:
        return round(float(value), 4)
    except Exception:
        return None


def _style_header(ws, row: int = 1) -> None:
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _autosize(ws, max_width: int = 48) -> None:
    for col_idx, column in enumerate(ws.columns, start=1):
        width = 10
        for cell in column:
            value = "" if cell.value is None else str(cell.value)
            width = max(width, min(max_width, max(len(part) for part in value.splitlines() or [""])))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width + 2


def _freeze_filter(ws) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _summary_sheet(wb: Workbook, report: dict) -> None:
    ws = wb.active
    ws.title = "Summary"
    headers = [
        "System",
        "Metric",
        "P@5",
        "R@5",
        "F1@5",
        "Jaccard",
        "nDCG@5",
        "Weighted P@5",
        "Mean relevance",
        "MRR",
        "Query coverage",
        "Keyword coverage",
        "Elapsed (s)",
    ]
    ws.append(headers)
    for combo in report.get("combos") or []:
        agg = combo.get("aggregate") or {}
        cov = combo.get("coverage_summary") or {}
        ws.append([
            combo.get("model"),
            combo.get("metric"),
            _fmt_float(agg.get("precision_at_5_mean")),
            _fmt_float(agg.get("recall_at_5_mean")),
            _fmt_float(agg.get("f1_at_5_mean")),
            _fmt_float(agg.get("jaccard_mean")),
            _fmt_float(agg.get("ndcg_at_5_mean")),
            _fmt_float(agg.get("weighted_precision_at_5")),
            _fmt_float(agg.get("mean_relevance_at_5")),
            _fmt_float(agg.get("mrr_relevance_ge_2")),
            _fmt_float(cov.get("query_coverage_mean")),
            _fmt_float(cov.get("keyword_coverage_mean")),
            _fmt_float(combo.get("elapsed_s")),
        ])
    _style_header(ws)

    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Semantic Model Comparison"
    chart.y_axis.title = "Mean score"
    chart.x_axis.title = "System"
    data = Reference(ws, min_col=3, max_col=8, min_row=1, max_row=1 + len(report.get("combos") or []))
    cats = Reference(ws, min_col=1, min_row=2, max_row=1 + len(report.get("combos") or []))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 8
    chart.width = 18
    ws.add_chart(chart, "O2")

    ws["A7"] = "Methodology"
    ws["A7"].font = Font(bold=True)
    ws["A8"] = (
        "No-cache video representation. Same 15 videos, same VLM title+summary, "
        "same 116-keyword pool, same cosine similarity. Graded VLM reference "
        "uses relevance 3/2/1; binary metrics use relevance >= 2."
    )
    ws["A8"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A8:M10")

    _autosize(ws)


def _per_video_sheet(wb: Workbook, report: dict, graded: dict) -> None:
    ws = wb.create_sheet("Per Video Top5")
    combos = {c.get("model"): c for c in report.get("combos") or []}
    graded_items = {i.get("id"): i for i in graded.get("items") or []}
    headers = [
        "Video ID",
        "URL",
        "Title",
        "Topic",
        "Subtopic",
        "Graded VLM reference (keyword relevance)",
        "BERT+VLM top-5",
        "BERT nDCG@5",
        "Word2Vec+VLM top-5",
        "Word2Vec nDCG@5",
        "GloVe+VLM top-5",
        "GloVe nDCG@5",
    ]
    ws.append(headers)
    for q in report.get("queries") or []:
        vid = q.get("item_id")
        item = graded_items.get(vid, {})
        row = [
            vid,
            item.get("url"),
            item.get("title"),
            item.get("topic"),
            item.get("subtopic"),
            _join_graded(item.get("graded_reference")),
        ]
        for model in ("bert", "word2vec", "glove"):
            combo = combos.get(model) or {}
            rec = next((x for x in combo.get("items") or [] if x.get("item_id") == vid), {})
            row.extend([
                _join_keywords(rec.get("top_keywords")),
                _fmt_float((rec.get("graded_metrics") or {}).get("ndcg_at_5")),
            ])
        ws.append(row)
    _style_header(ws)
    _freeze_filter(ws)
    _autosize(ws, max_width=62)


def _long_sheet(wb: Workbook, report: dict, graded: dict) -> None:
    ws = wb.create_sheet("Long Results")
    graded_items = {i.get("id"): i for i in graded.get("items") or []}
    headers = [
        "Video ID",
        "Title",
        "Topic",
        "Subtopic",
        "System",
        "Top 1",
        "Top 2",
        "Top 3",
        "Top 4",
        "Top 5",
        "P@5",
        "R@5",
        "F1@5",
        "Jaccard",
        "nDCG@5",
        "Weighted P@5",
        "Mean relevance",
        "MRR",
    ]
    ws.append(headers)
    for combo in report.get("combos") or []:
        model = combo.get("model")
        for rec in combo.get("items") or []:
            vid = rec.get("item_id")
            item = graded_items.get(vid, {})
            top = list(rec.get("top_keywords") or [])[:5]
            gm = rec.get("graded_metrics") or {}
            ws.append([
                vid,
                item.get("title"),
                item.get("topic"),
                item.get("subtopic"),
                model,
                *(top + [""] * (5 - len(top))),
                _fmt_float(rec.get("precision_at_5")),
                _fmt_float(rec.get("recall_at_5")),
                _fmt_float(rec.get("f1_at_5")),
                _fmt_float(rec.get("jaccard")),
                _fmt_float(gm.get("ndcg_at_5")),
                _fmt_float(gm.get("weighted_precision_at_5")),
                _fmt_float(gm.get("mean_relevance_at_5")),
                _fmt_float(gm.get("mrr_relevance_ge_2")),
            ])
    _style_header(ws)
    _freeze_filter(ws)
    _autosize(ws)


def _reference_sheet(wb: Workbook, graded: dict) -> None:
    ws = wb.create_sheet("Graded Reference")
    headers = ["Video ID", "Title", "Topic", "Subtopic", "Keyword", "Relevance", "Rationale"]
    ws.append(headers)
    for item in graded.get("items") or []:
        for ref in item.get("graded_reference") or []:
            ws.append([
                item.get("id"),
                item.get("title"),
                item.get("topic"),
                item.get("subtopic"),
                ref.get("keyword"),
                ref.get("relevance"),
                ref.get("rationale"),
            ])
    _style_header(ws)
    _freeze_filter(ws)
    _autosize(ws)


def main(argv=None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--report", default="reports/semantic_model_comparison_15videos_graded_nocache.json")
    parser.add_argument("--graded", default="data/gold/videos_15_graded_vlm_reference_nocache.json")
    parser.add_argument("--out", default="reports/semantic_model_comparison_15videos_graded_nocache.xlsx")
    args = parser.parse_args(argv)

    report = _load(_abs(args.report))
    graded = _load(_abs(args.graded))

    wb = Workbook()
    _summary_sheet(wb, report)
    _per_video_sheet(wb, report, graded)
    _long_sheet(wb, report, graded)
    _reference_sheet(wb, graded)

    out = _abs(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(out)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
